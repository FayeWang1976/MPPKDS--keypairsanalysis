#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Dec  7 18:32:08 2022

@author: wangzhehan
"""
import openpyxl
import numpy as np
import math

if __name__ == '__main__':
    
    # read key_pairs sheet
    split_key_list = openpyxl.load_workbook('split_key_list_for_matlab.xlsx')

    sheet = split_key_list['Sheet1']

    num_pairs = 10

    # read input
    input = [[] for j in range(num_pairs)]
    i = -1
    for row in sheet.iter_rows(min_row=2, max_row=num_pairs+1, min_col=1, max_col=24):
        i += 1
        for cell in row:
             input[i].append(cell.value)
        
    # read output
    output = [[] for j in range(num_pairs)]
    i = -1
    for row in sheet.iter_rows(min_row=2, max_row=num_pairs+1, min_col=25, max_col=48):
        i += 1
        for cell in row:
            output[i].append(cell.value)


    # calculate meanIn meanOut
    meanIn = []
    meanOut = []
    for j in range(num_pairs):
        meanIn.append(np.mean(input[j][0:20]))
        meanOut.append(np.mean(output[j][0:16]))
    
    meanIn[0] = np.mean(input[0])
    meanOut[0] = np.mean(output[0])

    # in-delt out-delt
    in_delt = [[] for j in range(num_pairs)]
    out_delt = [[] for j in range(num_pairs)]
    for j in range(num_pairs):
        for k in range(len(input[j])):
            in_delt[j].append(input[j][k]-meanIn[j])
            out_delt[j].append(output[j][k]-meanOut[j])


    # ups downs
    ups = [[] for j in range(num_pairs)]
    downs = [[] for j in range(num_pairs)]
    for j in range(num_pairs):
        for k in range(len(input[j])):
            ups[j].append(max(0, in_delt[j][k] * out_delt[j][k]))
            downs[j].append(max(0, -in_delt[j][k] * out_delt[j][k]))


    # semi_corr
    inStd = []
    outStd = []
    for j in range(num_pairs):
        inStd.append(np.std(input[j],ddof = 1))
        outStd.append(np.std(output[j],ddof = 1))

    up = []
    down = []
    for j in range(num_pairs):
        up.append(math.sqrt(np.mean(ups[j]))/math.sqrt(inStd[j]*outStd[j]))
        down.append(math.sqrt(np.mean(downs[j]))/math.sqrt(inStd[j]*outStd[j]))

    semi_corr_up = np.mean(up)
    semi_corr_down = np.mean(down)


    
    
